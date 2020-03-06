import psycopg2
from openpyxl import load_workbook
import re

def connect_to_db():
    conn = psycopg2.connect(dbname='d6vjl3au6tld96', user='exmeqvroilujbg',
                            password='2dbe50b6c122dfabebc0f19ad65d8d6e180f13927943103314036738f4ffad22',
                            host='ec2-54-228-181-43.eu-west-1.compute.amazonaws.com')

    cursor = conn.cursor()
    return conn, cursor


def create():
    conn, cursor = connect_to_db()

    cursor.execute("CREATE TABLE library (id serial PRIMARY KEY, title varchar, annotations varchar, author varchar, "
                   "EGEDirection varchar, EssayDirection varchar, OGEDirection varchar, is_have_illustration bool, "
                   "number_of_copies integer, number_of_pages integer, section varchar, year_of_publish integer);")

    wb_example = load_workbook('./table.xlsx')
    ws_example = wb_example.active

    for row in ws_example.values:
        if row[6] == '=TRUE()':
            cursor.execute("INSERT INTO library (title, annotations, author, EGEDirection, EssayDirection,OGEDirection,"
                           "is_have_illustration, number_of_copies, number_of_pages, section, year_of_publish) "
                           "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                           (row[0], row[1], row[2], row[3], row[4], row[5], True, row[7], row[8], row[9], row[10]))
        else:
            cursor.execute("INSERT INTO library (title, annotations, author, EGEDirection, EssayDirection,OGEDirection,"
                           "is_have_illustration, number_of_copies, number_of_pages, section, year_of_publish) "
                           "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                           (row[0], row[1], row[2], row[3], row[4], row[5], False, row[7], row[8], row[9], row[10]))

    conn.commit()
    cursor.close()
    conn.close()


def update_from_excel():
    conn, cur = connect_to_db()
    wb_example = load_workbook('./src/excel/table2020.xlsx')
    ws_example = wb_example.active
    cur.execute('SELECT * FROM library;')
    books = sorted(list(cur.fetchall()), key=lambda a: ''.join(a[3].split(' ')))
    books_ex = sorted(list(ws_example.values), key=lambda a: ''.join(a[2].split(' ')))

    # Сорри за этот говнокод, это пиздец

    for i in range(len(books)):
        k = list(books_ex[i])
        k[2] = books[i][3]
        books_ex[i] = tuple(k)

    cnt = 0
    cnt_p = 0
    names = ''
    for row in books_ex:
        print(row)
        cur.execute("UPDATE library SET EssayDirection = '{essay}' WHERE (year_of_publish = '{year}' AND number_of_pages = '{pages}' AND author = '{author}') OR (title = '{title}' AND author = '{author}');"
                    .format(essay=row[0] if row[0] is not None else 'Нет', title=row[1], year=row[4], pages=row[5], author=row[2]))

        cur.execute("SELECT title FROM library WHERE (year_of_publish = '{year}' AND number_of_pages = '{pages}' AND author = '{author}') OR (title = '{title}' AND author = '{author}');"
                    .format(title=row[1], year=row[4], pages=row[5], author=row[2]))
        a = cur.fetchall()

        if len(a) > 1:
            cnt_p += 1

        print(a)
        cnt += 0 if len(a) else 1
        names += '' if len(a) else row[1] + '; '

    print('Num times? when founded more then 1 book:', cnt_p)
    print('Number not found:', cnt)
    print('No founded books:\n', names)
    if not cnt_p and not cnt:
        conn.commit()
    cur.close()
    conn.close()


def find(name):
    conn, cur = connect_to_db()
    cur.execute('SELECT * FROM library;')

    books = []

    for row in cur.fetchall():
        if name.lower() in row[1].lower() or name.lower() in row[3].lower():
            books.append(row[1:])

    cur.close()
    conn.close()
    return books


def get_book_on_direction(what_direction, name_direction):
    conn, cur = connect_to_db()
    cur.execute('SELECT * FROM library;')
    books = []
    split = False
    if what_direction == 'Section':
        number_of_cell = 10
    elif what_direction == 'OGEDirection':
        number_of_cell = 6
    elif what_direction == 'EGEDirection':
        number_of_cell = 4
    elif what_direction == 'EssayDirection':
        number_of_cell = 5
        split = True
    else:
        return [Exception]

    for row in cur.fetchall():
        if split:
            if row[number_of_cell] is not None and name_direction in row[number_of_cell].split('/'):
                books.append(row[1:])
        else:
            if name_direction == row[number_of_cell]:
                books.append(row[1:])

    cur.close()
    conn.close()
    return books


def get_all_books():
    conn, cursor = connect_to_db()

    cursor.execute('SELECT * FROM library;')
    books = []
    for row in cursor.fetchall():
        books.append(row[1:])

    cursor.close()
    conn.close()

    return books


def drop_table(name):
    conn, cur = connect_to_db()

    cur.execute(f'DROP TABLE "{name}";')

    conn.commit()
    conn.close()
