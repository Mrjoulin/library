from openpyxl import load_workbook
import psycopg2
conn = psycopg2.connect(dbname='d6vjl3au6tld96', user='exmeqvroilujbg',
                        password='2dbe50b6c122dfabebc0f19ad65d8d6e180f13927943103314036738f4ffad22',
                        host='ec2-54-228-181-43.eu-west-1.compute.amazonaws.com')

cursor = conn.cursor()
cursor.execute('SELECT * FROM book;')

wb_example = load_workbook('./table.xlsx')
ws_example = wb_example.active

wb_input = load_workbook('./библиотека.xlsx')
ws_input = wb_input.active
# EGE - 3
# Essay - 4
# OGE - 5
OGEDirections = [
    "Преданность",
    "Человечность",
    "Любовь",
    "Настоящее искусство",
    "Сострадание",
    "Доброта",
    "Самовоспитание",
    "Подвиг",
    "Совесть"
]

EGEDirecions = [
    "Человек и природа",
    "Проблемы семьи",
    "Проблемы, связанные с отрицательными качествами личности",
    "Проблемы, связанные с положительными нравственными качествами личности",
    "Проблемы, связанные с ролью искусства и литературы в жизни человека",
    "Человек и общество",
    "Военная проблематика",
    "Социальной ответственности ученых за их изобретения",
    "Проблема одиночества",
    "Человек и государственная власть (политическая)"
]

EssayDirections = [
    "«Верность и измена»",
    "«Равнодушие и отзывчивость»",
    "«Цели и средства»",
    "«Смелость и трусость»",
    "«Человек и общество»"
]
Directions = [EGEDirecions, EssayDirections, OGEDirections]
rowNumber = 1

alph = 'abcdefghijklmnopqrstuvwxyz'
for row in ws_example.values:
    print(row[6], bool(row[6]))

wb_example.save('./table.xlsx')
cursor.close()
conn.close()
